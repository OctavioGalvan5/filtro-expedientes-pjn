# -*- coding: utf-8 -*-
"""
PJN Scraper - Módulo de Login y Navegación
Automación con Selenium y Chrome
"""

import os
import re
import sys
import time
import traceback
import requests
import pdfplumber
import openpyxl
from datetime import datetime
import database as db
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

FRASE_BUSCADA = "apoderado de la Caja de Seguridad Social para Abogados de la Provincia de Salta"
ID_NUEVA_CONSULTA = "j_idt24:menuNavigation:j_idt36:menuNuevaConsulta"

# Flag de parada para la GUI; se pone en True desde gui.py
_stop_requested = False


def inicializar_navegador(headless=False):
    from webdriver_manager.core.os_manager import ChromeType

    options = Options()
    en_docker = os.environ.get("DOCKER_ENV") == "1"

    if en_docker or headless:
        log("[INFO] Modo HEADLESS activado" + (" (Docker)" if en_docker else ""))
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
    else:
        log("[INFO] Modo NORMAL (con ventana visible) activado")
        options.add_argument("--start-maximized")

    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    if en_docker:
        # ChromeDriverManager detecta la versión de Chromium instalada y descarga
        # el driver correcto, evitando el mismatch de versiones con /usr/bin/chromedriver
        service = Service(
            ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install(),
            service_args=["--verbose"],
            log_output="/tmp/chromedriver.log",
        )
        driver = webdriver.Chrome(service=service, options=options)
    else:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    driver.implicitly_wait(10)
    return driver


def extraer_texto_pdf(url_pdf, cookies):
    nombre_temporal = f"temp_doc_{abs(hash(url_pdf)) % 100000}.pdf"
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        resp = requests.get(url_pdf, headers=headers, cookies=cookies, timeout=60, stream=True)
        resp.raise_for_status()

        with open(nombre_temporal, 'wb') as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        texto_completo = ""
        with pdfplumber.open(nombre_temporal) as pdf:
            for pagina in pdf.pages:
                texto_pagina = pagina.extract_text()
                if texto_pagina:
                    texto_completo += texto_pagina + "\n"

        return texto_completo.strip() if texto_completo.strip() else None
    except Exception as e:
        print(f"      [Error PDF] No se pudo procesar: {e}")
        return None
    finally:
        if os.path.exists(nombre_temporal):
            try:
                os.remove(nombre_temporal)
            except Exception:
                pass


def _procesar_filas(driver, filas_datos, pagina_num):
    """Descarga cada documento, extrae texto y busca la frase. Retorna hallazgos."""

    # Paso 1: extraer todos los datos del DOM antes de hacer cualquier descarga.
    # Así los elementos no pueden volverse stale durante las operaciones lentas.
    filas_info = []
    for fila in filas_datos:
        try:
            celdas = fila.find_elements(By.TAG_NAME, "td")
            if len(celdas) < 6:
                continue
            viewer_url = None
            for a in celdas[0].find_elements(By.TAG_NAME, "a"):
                href = a.get_attribute("href") or ""
                if "viewer.seam" in href and "download=true" not in href:
                    viewer_url = href
                    break
            filas_info.append({
                "oficina":     celdas[1].text.strip(),
                "fecha":       celdas[2].text.strip(),
                "tipo":        celdas[3].text.strip(),
                "descripcion": celdas[4].text.strip(),
                "foja":        celdas[5].text.strip(),
                "viewer_url":  viewer_url,
            })
        except StaleElementReferenceException:
            log(f"[Advertencia] Fila obsoleta en pág.{pagina_num}, se omite.")

    # Paso 2: procesar cada fila con los datos ya en memoria (sin tocar el DOM).
    hallazgos = []
    cookies = {c['name']: c['value'] for c in driver.get_cookies()}

    for idx, info in enumerate(filas_info, 1):
        viewer_url = info["viewer_url"]
        contenido = "No contiene documento adjunto para ver"
        frase_encontrada = False

        if viewer_url:
            download_url = viewer_url + "&download=true"
            print(f"      [Descarga] Pág.{pagina_num} fila {idx}: descargando documento...")
            texto = extraer_texto_pdf(download_url, cookies)
            if texto:
                contenido = texto
                contenido_normalizado = " ".join(contenido.split())
                frase_encontrada = FRASE_BUSCADA.lower() in contenido_normalizado.lower()
            else:
                contenido = "No se pudo extraer el texto del documento"

        print(f"\n--- Pág.{pagina_num} | Fila {idx} ---")
        print(f"Fecha: {info['fecha']}")
        print(f"Oficina: {info['oficina']}")
        print(f"Tipo: {info['tipo']}")
        print(f"Descripcion: {info['descripcion']}")
        print(f"Foja: {info['foja']}")
        if frase_encontrada:
            print("*** FRASE ENCONTRADA EN ESTE DOCUMENTO — DETENIENDO PROCESO ***")
            hallazgos.append({"seccion": "", "pagina": pagina_num, "fila": idx,
                               "fecha": info["fecha"], "tipo": info["tipo"],
                               "descripcion": info["descripcion"]})
            print(f"Contenido:\n{contenido}")
            print("-" * 60)
            break
        print(f"Contenido:\n{contenido}")
        print("-" * 60)

    return hallazgos


def _procesar_tabla_completa(driver, tabla_id, seccion_nombre, buscar_frase=True):
    """
    Único pase por todas las páginas de la tabla.
    Para cada fila:
      - Siempre: extrae fecha (fecha_ultima) y detecta demanda (url/fecha/detalle).
      - Si buscar_frase=True y la frase aún no fue encontrada: descarga el PDF y busca.
        Una vez encontrada la frase, deja de descargar PDFs pero continúa las páginas
        para completar fecha_ultima y demanda más antigua.
    Retorna (hallazgos, fecha_ultima, url_demanda, fecha_demanda, detalle_demanda).
    """
    pagina_num    = 1
    hallazgos     = []
    frase_encontrada = False
    fecha_ultima  = None
    _demandas     = []  # acumula todas las filas que coinciden con demanda

    while True:
        log(f"[{seccion_nombre}] Procesando página {pagina_num}...")

        filas_datos = None
        for intento in range(3):
            try:
                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located((By.ID, tabla_id))
                )
                filas_datos = driver.find_element(By.ID, tabla_id).find_elements(By.TAG_NAME, "tr")[1:]
                break
            except StaleElementReferenceException:
                if intento == 2:
                    raise
                log(f"[{seccion_nombre}] Tabla obsoleta, reintentando ({intento + 1}/3)...")
                time.sleep(1)

        log(f"[{seccion_nombre}] {len(filas_datos)} movimientos en esta página.")

        # Paso 1: volcar el DOM a memoria antes de cualquier descarga
        filas_info = []
        for fila in filas_datos:
            try:
                celdas = fila.find_elements(By.TAG_NAME, "td")
                if len(celdas) < 5:
                    continue
                raw_fecha = celdas[2].text.strip()
                m = re.search(r'\d{1,2}/\d{1,2}/\d{4}', raw_fecha)
                viewer_url = None
                for a in celdas[0].find_elements(By.TAG_NAME, "a"):
                    href = a.get_attribute("href") or ""
                    if "viewer.seam" in href and "download=true" not in href:
                        viewer_url = href
                        break
                filas_info.append({
                    "raw_fecha":   raw_fecha,
                    "fecha_m":     m,
                    "oficina":     celdas[1].text.strip() if len(celdas) > 1 else "",
                    "tipo":        celdas[3].text.strip(),
                    "detalle_raw": celdas[4].text.strip(),
                    "foja":        celdas[5].text.strip() if len(celdas) > 5 else "",
                    "viewer_url":  viewer_url,
                })
            except StaleElementReferenceException:
                log(f"[Advertencia] Fila obsoleta en pág.{pagina_num}, se omite.")

        # Paso 2: procesar con datos en memoria
        cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        for idx, info in enumerate(filas_info, 1):
            m         = info["fecha_m"]
            tipo_up   = info["tipo"].upper()
            det_up    = info["detalle_raw"].upper()

            # Fecha más antigua (siempre)
            if m:
                fecha_ultima = m.group(0)

            # Colectar TODAS las filas de demanda (todas las partes)
            if ("ESCRITO" in tipo_up and "DEMANDA" in det_up
                    and "DESISTE" not in det_up and info["viewer_url"]):
                _pm = re.search(r'\(Parte\s+(\d+)\s+de\s+\d+\)', info["detalle_raw"], re.IGNORECASE)
                _demandas.append({
                    "url":    info["viewer_url"],
                    "fecha":  m.group(0) if m else None,
                    "detalle": info["detalle_raw"],
                    "parte":  int(_pm.group(1)) if _pm else 0,
                })

            # Descarga PDF para frase (solo si aún no fue encontrada)
            contenido = "No contiene documento adjunto para ver"
            if buscar_frase and not frase_encontrada and info["viewer_url"]:
                print(f"      [Descarga] Pág.{pagina_num} fila {idx}: descargando documento...")
                texto = extraer_texto_pdf(info["viewer_url"] + "&download=true", cookies)
                if texto:
                    contenido = texto
                    if FRASE_BUSCADA.lower() in " ".join(contenido.split()).lower():
                        frase_encontrada = True
                        hallazgos.append({
                            "seccion":     seccion_nombre,
                            "pagina":      pagina_num,
                            "fila":        idx,
                            "fecha":       info["raw_fecha"],
                            "tipo":        info["tipo"],
                            "descripcion": info["detalle_raw"],
                        })
                        print("*** FRASE ENCONTRADA EN ESTE DOCUMENTO ***")
                else:
                    contenido = "No se pudo extraer el texto del documento"

            print(f"\n--- Pág.{pagina_num} | Fila {idx} ---")
            print(f"Fecha: {info['raw_fecha']}")
            print(f"Oficina: {info['oficina']}")
            print(f"Tipo: {info['tipo']}")
            print(f"Descripcion: {info['detalle_raw']}")
            print(f"Foja: {info['foja']}")
            print(f"Contenido:\n{contenido}")
            print("-" * 60)

        btn_sig = _encontrar_boton_siguiente(driver)
        if not btn_sig:
            log(f"[{seccion_nombre}] Última página. Total: {pagina_num} página(s).")
            break

        log(f"[{seccion_nombre}] Avanzando a página {pagina_num + 1}...")
        primera_fila = filas_datos[0] if filas_datos else None
        btn_sig.click()
        if primera_fila:
            WebDriverWait(driver, 20).until(EC.staleness_of(primera_fila))
            time.sleep(0.5)   # dar margen a que el AJAX termine de inyectar la nueva tabla
        else:
            time.sleep(2)
        pagina_num += 1

    # Consolidar todas las partes encontradas
    if _demandas:
        _dem_ord      = sorted(_demandas, key=lambda d: d["parte"])  # Parte 1 primero
        url_demanda   = "|".join(d["url"] for d in _dem_ord)
        fecha_demanda = _dem_ord[-1]["fecha"]  # la más antigua
        detalle_demanda = re.sub(
            r'\s*\(Parte\s+\d+\s+de\s+\d+\)', '', _dem_ord[0]["detalle"], flags=re.IGNORECASE
        ).strip()
    else:
        url_demanda = fecha_demanda = detalle_demanda = None

    return hallazgos, fecha_ultima, url_demanda, fecha_demanda, detalle_demanda


def extraer_fecha_ultima_fila(driver, tabla_id):
    """Navega hasta la última página de la tabla y devuelve la fecha de la última fila."""
    try:
        while _boton_siguiente_habilitado(driver):
            WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.ID, tabla_id))
            )
            filas = driver.find_element(By.ID, tabla_id).find_elements(By.TAG_NAME, "tr")
            primera_fila = filas[1] if len(filas) > 1 else None
            driver.find_element(By.XPATH, "//span[@title='Siguiente']").click()
            if primera_fila:
                WebDriverWait(driver, 15).until(EC.staleness_of(primera_fila))

        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, tabla_id))
        )
        filas = driver.find_element(By.ID, tabla_id).find_elements(By.TAG_NAME, "tr")[1:]
        if not filas:
            return None
        celdas = filas[-1].find_elements(By.TAG_NAME, "td")
        if len(celdas) > 2:
            for span in celdas[2].find_elements(By.TAG_NAME, "span"):
                texto = span.text.strip()
                if texto:
                    return texto
        return None
    except Exception as e:
        log(f"[FechaInicio] Error extrayendo fecha: {e}")
        return None


def extraer_datos_inicio(driver, tabla_id):
    """
    Recorre todas las páginas de la tabla en un único pase.
    Devuelve (fecha_ultima, url_demanda, fecha_demanda):
      - fecha_ultima:  fecha de la última fila de la última página (más antigua)
      - url_demanda:   viewer URL del primer ESCRITO con DEMANDA (excluye DESISTE)
      - fecha_demanda: fecha de esa misma fila de demanda
    """
    fecha_ultima = None
    _demandas    = []  # acumula todas las filas que coinciden con demanda

    while True:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, tabla_id))
        )
        filas = driver.find_element(By.ID, tabla_id).find_elements(By.TAG_NAME, "tr")[1:]
        if not filas:
            break

        for fila in filas:
            try:
                celdas = fila.find_elements(By.TAG_NAME, "td")
                if len(celdas) < 5:
                    continue

                raw_fecha = celdas[2].text.strip()
                m = re.search(r'\d{1,2}/\d{1,2}/\d{4}', raw_fecha)
                if m:
                    fecha_ultima = m.group(0)

                tipo        = celdas[3].text.strip().upper()
                detalle_raw = celdas[4].text.strip()
                detalle     = detalle_raw.upper()
                if "ESCRITO" in tipo and "DEMANDA" in detalle and "DESISTE" not in detalle:
                    for a in celdas[0].find_elements(By.TAG_NAME, "a"):
                        href = a.get_attribute("href") or ""
                        if "viewer.seam" in href and "download=true" not in href:
                            _pm = re.search(r'\(Parte\s+(\d+)\s+de\s+\d+\)', detalle_raw, re.IGNORECASE)
                            _demandas.append({
                                "url":    href,
                                "fecha":  m.group(0) if m else None,
                                "detalle": detalle_raw,
                                "parte":  int(_pm.group(1)) if _pm else 0,
                            })
                            break
            except StaleElementReferenceException:
                pass

        btn_sig = _encontrar_boton_siguiente(driver)
        if not btn_sig:
            break

        primera_fila = filas[0]
        btn_sig.click()
        WebDriverWait(driver, 15).until(EC.staleness_of(primera_fila))

    # Consolidar todas las partes encontradas
    if _demandas:
        _dem_ord        = sorted(_demandas, key=lambda d: d["parte"])  # Parte 1 primero
        url_demanda     = "|".join(d["url"] for d in _dem_ord)
        fecha_demanda   = _dem_ord[-1]["fecha"]  # la más antigua
        detalle_demanda = re.sub(
            r'\s*\(Parte\s+\d+\s+de\s+\d+\)', '', _dem_ord[0]["detalle"], flags=re.IGNORECASE
        ).strip()
    else:
        url_demanda = fecha_demanda = detalle_demanda = None

    return fecha_ultima, url_demanda, fecha_demanda, detalle_demanda


def extraer_datos_inicio_expediente(driver, num, anio):
    """
    Navega a 'Nueva Consulta', busca el expediente y extrae en un único pase:
      - fecha_inicio (fecha más antigua de la última página)
      - url_demanda, fecha_demanda, detalle_demanda (del ESCRITO+DEMANDA más antiguo)
    Busca también en históricas. Puede llamarse desde cualquier página del portal.
    Devuelve un 4-tuple (fecha_inicio, url_demanda, fecha_demanda, detalle_demanda).
    """
    # Volver al formulario — XPath flexible (los j_idt son dinámicos) + JS click
    # para esquivar overlays que quedan tras procesar históricas.
    elem = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(@id,'menuNuevaConsulta')]"))
    )
    driver.execute_script("arguments[0].click();", elem)
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "formPublica:camaraNumAni"))
    )

    Select(driver.find_element(By.ID, "formPublica:camaraNumAni")).select_by_value("24")
    campo = driver.find_element(By.ID, "formPublica:numero")
    campo.clear()
    campo.send_keys(num)
    campo = driver.find_element(By.ID, "formPublica:anio")
    campo.clear()
    campo.send_keys(anio)
    driver.find_element(By.ID, "formPublica:buscarPorNumeroButton").click()

    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "expediente:action-table"))
    )

    fecha, url_demanda, fecha_demanda, detalle_demanda = extraer_datos_inicio(
        driver, "expediente:action-table"
    )
    log(f"[Inicio] Última fecha: {fecha} | Demanda: {'sí' if url_demanda else 'no'}")

    try:
        div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "expediente:btnActuacionesHistoricas"))
        )
        div.find_element(By.TAG_NAME, "a").click()
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "expediente:action-historic-table"))
        )
        f2, u2, fd2, dd2 = extraer_datos_inicio(driver, "expediente:action-historic-table")
        if f2:
            fecha = f2
            log(f"[Historicas] Fecha más antigua: {fecha}")
        if u2:
            url_demanda, fecha_demanda, detalle_demanda = u2, fd2, dd2
    except TimeoutException:
        log("[Historicas] Sin registros históricos.")
    except Exception:
        log(f"[Historicas] Error inesperado: {traceback.format_exc()}")

    return fecha, url_demanda, fecha_demanda, detalle_demanda


def _encontrar_boton_siguiente(driver):
    """
    Devuelve el elemento clickeable del botón Siguiente, o None si no hay más páginas.
    Soporta dos formatos del portal PJN:
      - Formato A (actuaciones): <span title='Siguiente'>
      - Formato B (históricas):  <a class='padding-pagination'>...<span>Siguiente</span></a>
    """
    # Formato A
    try:
        span = driver.find_element(By.XPATH, "//span[@title='Siguiente']")
        if span.is_displayed():
            elem = span
            for _ in range(3):
                parent = elem.find_element(By.XPATH, "..")
                if "disabled" in (parent.get_attribute("class") or "").lower():
                    break
                elem = parent
            else:
                return span
    except Exception:
        pass

    # Formato B
    try:
        a = driver.find_element(
            By.XPATH, "//a[contains(@class,'padding-pagination') and contains(.,'Siguiente')]"
        )
        if a.is_displayed():
            parent = a.find_element(By.XPATH, "..")
            if "disabled" not in (parent.get_attribute("class") or "").lower():
                return a
    except Exception:
        pass

    return None


def _boton_siguiente_habilitado(driver):
    return _encontrar_boton_siguiente(driver) is not None


def extraer_jurisdiccion_dependencia(driver):
    """Extrae jurisdicción, juzgado y secretaría del encabezado del expediente."""
    jurisdiccion = ""
    juzgado = ""
    secretaria = ""
    try:
        el = driver.find_element(By.XPATH, "//span[contains(@id,'detailCamera')]")
        jurisdiccion = el.text.strip()
    except Exception:
        pass
    try:
        el = driver.find_element(By.XPATH, "//span[contains(@id,'detailDependencia')]")
        dependencia = el.text.strip()
        if " - " in dependencia:
            partes = dependencia.split(" - ", 1)
            juzgado    = partes[0].strip()
            secretaria = partes[1].strip()
        else:
            juzgado = dependencia
    except Exception:
        pass
    return jurisdiccion, juzgado, secretaria


def extraer_intervinientes(driver):
    """
    Abre el tab Intervinientes, extrae partes y sus abogados, y vuelve al tab Actuaciones.
    Retorna lista de dicts: [{tipo, nombre, abogados:[{nombre, tomo_folio, cuit}]}]
    """
    participantes = []
    try:
        log("[Intervinientes] Abriendo tab...")
        tab = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                "//span[contains(@class,'rf-tab-lbl') and normalize-space()='Intervinientes']"
            ))
        )
        tab.click()
        time.sleep(1)

        tabla = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "expediente:participantsTable"))
        )

        tbodies = tabla.find_elements(By.TAG_NAME, "tbody")
        participante_actual = None

        for tbody in tbodies:
            cls = tbody.get_attribute("class") or ""
            if "rf-dt-b" in cls:
                for tr in tbody.find_elements(By.TAG_NAME, "tr"):
                    celdas = tr.find_elements(By.TAG_NAME, "td")
                    if len(celdas) < 2:
                        continue
                    tipo   = celdas[0].text.strip()
                    nombre = celdas[1].text.strip()
                    if nombre:
                        participante_actual = {"tipo": tipo, "nombre": nombre, "abogados": []}
                        participantes.append(participante_actual)
            elif "rf-cst" in cls and participante_actual is not None:
                for tr in tbody.find_elements(By.TAG_NAME, "tr"):
                    celdas = tr.find_elements(By.TAG_NAME, "td")
                    if not celdas:
                        continue
                    # col 0 = rol (LETRADO APODERADO, etc.)
                    # col 1 = nombre del abogado
                    # col 2 = tomo/folio
                    # col 3 = CUIT
                    nombre_ab  = celdas[1].text.strip() if len(celdas) > 1 else ""
                    tomo_folio = celdas[2].text.strip() if len(celdas) > 2 else ""
                    cuit       = celdas[3].text.strip() if len(celdas) > 3 else ""
                    if nombre_ab:
                        participante_actual["abogados"].append({
                            "nombre": nombre_ab,
                            "tomo_folio": tomo_folio,
                            "cuit": cuit,
                        })

        log(f"[Intervinientes] {len(participantes)} participante(s) extraído(s).")
    except Exception as e:
        log(f"[Intervinientes] No se pudo extraer: {e}")

    try:
        driver.find_element(By.XPATH,
            "//span[contains(@class,'rf-tab-lbl') and normalize-space()='Actuaciones']"
        ).click()
        time.sleep(1)
        # Solo esperar la tabla si el expediente tiene actuaciones
        if driver.find_elements(By.ID, "expediente:action-table"):
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.ID, "expediente:action-table"))
            )
    except Exception as e:
        log(f"[Intervinientes] Error volviendo a Actuaciones: {e}")

    return participantes


def procesar_tabla_paginada(driver, tabla_id, seccion_nombre):
    """Itera todas las páginas de una tabla y procesa cada fila. Retorna hallazgos."""
    pagina_num = 1
    todos_hallazgos = []

    while True:
        log(f"[{seccion_nombre}] Procesando página {pagina_num}...")

        # Reintentar hasta 3 veces si la tabla se vuelve stale justo al leerla
        filas_datos = None
        for intento in range(3):
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, tabla_id))
                )
                tabla = driver.find_element(By.ID, tabla_id)
                filas_datos = tabla.find_elements(By.TAG_NAME, "tr")[1:]
                break
            except StaleElementReferenceException:
                if intento == 2:
                    raise
                log(f"[{seccion_nombre}] Tabla obsoleta, reintentando ({intento + 1}/3)...")
                time.sleep(1)

        log(f"[{seccion_nombre}] {len(filas_datos)} movimientos en esta página.")

        hallazgos_pagina = _procesar_filas(driver, filas_datos, pagina_num)
        for h in hallazgos_pagina:
            h["seccion"] = seccion_nombre
        todos_hallazgos.extend(hallazgos_pagina)

        if todos_hallazgos:
            break

        btn_sig = _encontrar_boton_siguiente(driver)
        if not btn_sig:
            log(f"[{seccion_nombre}] Última página. Total: {pagina_num} página(s).")
            break

        log(f"[{seccion_nombre}] Avanzando a página {pagina_num + 1}...")
        primera_fila = filas_datos[0] if filas_datos else None
        btn_sig.click()

        if primera_fila:
            WebDriverWait(driver, 15).until(EC.staleness_of(primera_fila))
        else:
            time.sleep(1)

        pagina_num += 1

    return todos_hallazgos


def _login_y_abrir_formulario(driver, usuario, password):
    """Hace login y navega hasta el formulario de Nueva Consulta Pública."""
    log(f"[Navegando] Navegando al portal PJN...")
    driver.get("https://portalpjn.pjn.gov.ar/")

    log("[Esperando] Esperando redirección al Login (SSO)...")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "username")))

    log(f"[Login] Ingresando credenciales...")
    driver.find_element(By.ID, "username").send_keys(usuario)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "kc-login").click()

    log("[Esperando] Esperando portal de inicio...")
    xpath_consultas = "//span[contains(@class, 'MuiTypography-root') and text()='Consultas']"
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_consultas)))

    ventana_original = driver.current_window_handle
    driver.find_element(By.XPATH, xpath_consultas).click()

    WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > 1)
    for ventana in driver.window_handles:
        if ventana != ventana_original:
            driver.switch_to.window(ventana)
            break

    log("[Clic] Abriendo 'Nueva Consulta Pública'...")
    elem = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(@id,'menuNuevaConsulta')]"))
    )
    driver.execute_script("arguments[0].click();", elem)

    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "formPublica:camaraNumAni"))
    )
    log("[Exito] Formulario de consulta listo.\n")


def _buscar_y_procesar(driver, num_expediente, anio_expediente):
    """
    Busca el expediente y en un único pase por cada tabla:
      - Descarga PDFs y busca la frase
      - Extrae fecha_inicio, url_demanda, fecha_demanda, detalle_demanda
    Las históricas siempre se procesan (para obtener datos más antiguos);
    solo se busca la frase en ellas si aún no fue encontrada en la tabla principal.
    Retorna (encontrado, participantes, jurisdiccion, juzgado, secretaria,
             fecha_inicio, url_demanda, fecha_demanda, detalle_demanda).
    """
    log(f"[Consulta] Buscando expediente Nro {num_expediente} / Año {anio_expediente}...")

    select_camara = Select(driver.find_element(By.ID, "formPublica:camaraNumAni"))
    select_camara.select_by_value("24")

    input_numero = driver.find_element(By.ID, "formPublica:numero")
    input_numero.clear()
    input_numero.send_keys(num_expediente)

    input_anio = driver.find_element(By.ID, "formPublica:anio")
    input_anio.clear()
    input_anio.send_keys(anio_expediente)

    driver.find_element(By.ID, "formPublica:buscarPorNumeroButton").click()

    # Esperar que cargue la página del expediente: puede tener o no actuaciones.
    # El tab "Intervinientes" siempre aparece; action-table solo si hay actuaciones.
    try:
        WebDriverWait(driver, 30).until(
            lambda d: d.find_elements(By.ID, "expediente:action-table") or
                      d.find_elements(By.XPATH,
                          "//span[contains(@class,'rf-tab-lbl') and normalize-space()='Intervinientes']")
        )
    except TimeoutException:
        # Si seguimos en la página de búsqueda, el expediente no existe en PJN.
        if driver.find_elements(By.ID, "formPublica:buscarPorNumeroButton"):
            raise RuntimeError("Expediente no encontrado en PJN (sin resultados).")
        raise

    jurisdiccion, juzgado, secretaria = extraer_jurisdiccion_dependencia(driver)
    log(f"[Jurisdiccion] {jurisdiccion} | {juzgado} | {secretaria}")

    participantes = extraer_intervinientes(driver)

    tiene_actuaciones = bool(driver.find_elements(By.ID, "expediente:action-table"))
    if tiene_actuaciones:
        hallazgos, fecha_inicio, url_demanda, fecha_demanda, detalle_demanda = \
            _procesar_tabla_completa(driver, "expediente:action-table", "Actuaciones")
    else:
        log("[Actuaciones] Sin tabla de actuaciones en este expediente.")
        hallazgos, fecha_inicio, url_demanda, fecha_demanda, detalle_demanda = [], None, None, None, None

    # Históricas: siempre para fecha/demanda más antigua;
    # buscar frase solo si todavía no fue encontrada.
    print("\n[Historicas] Verificando historial de actuaciones...")
    try:
        div_historicas = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "expediente:btnActuacionesHistoricas"))
        )
        div_historicas.find_element(By.TAG_NAME, "a").click()
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "expediente:action-historic-table"))
        )
        log("[Historicas] Tabla encontrada. Procesando...")
        h2, fecha_hist, url_dem_hist, fd_hist, dd_hist = _procesar_tabla_completa(
            driver, "expediente:action-historic-table", "Historicas",
            buscar_frase=(not hallazgos),
        )
        hallazgos.extend(h2)
        if fecha_hist:
            fecha_inicio = fecha_hist
        if url_dem_hist:
            url_demanda, fecha_demanda, detalle_demanda = url_dem_hist, fd_hist, dd_hist
    except TimeoutException:
        log("[Historicas] Sin registros históricos.")
    except Exception:
        log("[Historicas] No se pudo acceder al historial.")

    return (len(hallazgos) > 0, participantes, jurisdiccion, juzgado, secretaria,
            fecha_inicio, url_demanda, fecha_demanda, detalle_demanda)


def ejecutar_flujo(usuario, password, headless=False):
    """Modo individual: un solo expediente hardcodeado."""
    driver = None
    try:
        driver = inicializar_navegador(headless=headless)
        _login_y_abrir_formulario(driver, usuario, password)

        num_expediente = "13000369"
        anio_expediente = "2006"

        encontrado, *_ = _buscar_y_procesar(driver, num_expediente, anio_expediente)

        print("\n" + "=" * 60)
        print("RESUMEN FINAL")
        print("=" * 60)
        print(f'Frase buscada: "{FRASE_BUSCADA}"')
        print()
        if encontrado:
            print("*** FRASE ENCONTRADA ***")
        else:
            print("La frase NO fue encontrada en ningún documento.")
        print("=" * 60)

    except Exception:
        log("[Error] Ocurrió un error:")
        traceback.print_exc()
    finally:
        if driver:
            log("[Cerrando] Cerrando el navegador...")
            driver.quit()


def _detectar_columnas(ws):
    """Detecta índices de columnas por encabezado (fila 1, case-insensitive)."""
    encabezados = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            encabezados[str(val).lower().strip()] = c

    def buscar_col(*opciones):
        for op in opciones:
            if op in encabezados:
                return encabezados[op]
        raise ValueError(f"No se encontró columna. Opciones buscadas: {opciones}")

    col_num      = buscar_col("numero", "número", "nro", "num")
    col_anio     = buscar_col("año", "anio", "year")
    col_caratula = buscar_col("caratula", "carátula")
    return col_num, col_anio, col_caratula, encabezados


def ejecutar_desde_excel(archivo_entrada, usuario, password,
                         headless=False, on_progreso=None, fuente=None):
    """
    Lee expedientes de expedientes.xlsx, procesa cada uno y guarda en PostgreSQL.
    La BD es la única fuente de verdad: no se genera ningún Excel de salida.
    on_progreso(actual, total): callback opcional llamado luego de cada expediente.
    fuente: nombre del archivo de origen (si None, se deriva del nombre del archivo).
    """
    global _stop_requested
    _stop_requested = False

    if fuente is None:
        fuente = os.path.basename(archivo_entrada)

    try:
        db.inicializar_db()

        wb_entrada = openpyxl.load_workbook(archivo_entrada)
        ws_entrada = wb_entrada.active
        col_num, col_anio, col_caratula, _ = _detectar_columnas(ws_entrada)
        total = ws_entrada.max_row - 1

        for row_idx in range(2, ws_entrada.max_row + 1):
            if _stop_requested:
                log("[Detenido] El usuario detuvo el proceso.")
                break

            num      = str(ws_entrada.cell(row_idx, col_num).value or "").strip()
            anio     = str(ws_entrada.cell(row_idx, col_anio).value or "").strip()
            caratula = str(ws_entrada.cell(row_idx, col_caratula).value or "").strip()

            fila_actual = row_idx - 1
            print(f"\n{'='*60}")
            log(f"[{fila_actual}/{total}] {caratula}")
            log(f"        Expediente: {num} / {anio}")
            print(f"{'='*60}")

            if not num or not anio:
                log("[Saltando] Fila sin numero o año.")
                continue

            if db.ya_fue_procesado(num, anio):
                log("[Saltando] Ya procesado previamente (en BD).")
                if on_progreso:
                    on_progreso(fila_actual, total)
                continue

            driver = None
            resultado = "Error"
            participantes = []
            jurisdiccion = juzgado = secretaria = ""
            fecha_inicio = url_demanda = fecha_demanda = detalle_demanda = None
            try:
                driver = inicializar_navegador(headless=headless)
                _login_y_abrir_formulario(driver, usuario, password)
                (encontrado, participantes, jurisdiccion, juzgado, secretaria,
                 fecha_inicio, url_demanda, fecha_demanda, detalle_demanda) = \
                    _buscar_y_procesar(driver, num, anio)
                resultado = "Si" if encontrado else "No"
                log(f"[Resultado] Caja se presenta: {resultado}")
                log(f"[Inicio] fecha_inicio={fecha_inicio}"
                    + (f" | fecha_demanda={fecha_demanda}" if fecha_demanda else "")
                    + f" | demanda={'sí' if url_demanda else 'no'}")
            except KeyboardInterrupt:
                raise
            except Exception:
                log(f"[Error] Fallo al procesar {num}/{anio}:\n{traceback.format_exc()}")
            finally:
                if driver:
                    driver.quit()

            try:
                db.guardar_expediente(
                    num, anio, caratula, resultado, participantes,
                    jurisdiccion, juzgado, secretaria, fuente,
                    fecha_inicio=fecha_inicio,
                    url_demanda=url_demanda or "NINGUNA",
                    fecha_demanda=fecha_demanda,
                    detalle_demanda=detalle_demanda,
                )
            except Exception:
                log(f"[Error] Fallo al guardar en BD {num}/{anio}:\n{traceback.format_exc()}")

            if on_progreso:
                on_progreso(fila_actual, total)

        log("[Exito] Proceso completado. Todos los resultados estan en PostgreSQL.")

    except KeyboardInterrupt:
        print("\n")
        log("[Interrumpido] Proceso detenido. Progreso guardado en la BD.")
        log("[Interrumpido] La proxima ejecucion retomara desde el ultimo expediente pendiente.")
    except Exception:
        log(f"[Error] Error en el proceso batch:\n{traceback.format_exc()}")


def ejecutar_lista(lista, usuario, password, headless=False, on_progreso=None):
    """
    Reintenta una lista de expedientes (dicts con numero, anio, caratula).
    No verifica ya_fue_procesado — siempre reprocesa y sobreescribe en BD.
    """
    global _stop_requested
    _stop_requested = False

    total = len(lista)

    try:
        db.inicializar_db()

        for idx, item in enumerate(lista, start=1):
            if _stop_requested:
                log("[Detenido] El usuario detuvo el proceso.")
                break

            num      = str(item.get("numero", "")).strip()
            anio     = str(item.get("anio", "")).strip()
            caratula = str(item.get("caratula") or "").strip()

            print(f"\n{'='*60}")
            log(f"[{idx}/{total}] {caratula}")
            log(f"        Expediente: {num} / {anio}")
            print(f"{'='*60}")

            if not num or not anio:
                log("[Saltando] Fila sin numero o año.")
                continue

            driver = None
            resultado = "Error"
            participantes = []
            jurisdiccion = juzgado = secretaria = ""
            fecha_inicio = url_demanda = fecha_demanda = detalle_demanda = None
            try:
                driver = inicializar_navegador(headless=headless)
                _login_y_abrir_formulario(driver, usuario, password)
                (encontrado, participantes, jurisdiccion, juzgado, secretaria,
                 fecha_inicio, url_demanda, fecha_demanda, detalle_demanda) = \
                    _buscar_y_procesar(driver, num, anio)
                resultado = "Si" if encontrado else "No"
                log(f"[Resultado] Caja se presenta: {resultado}")
                log(f"[Inicio] fecha_inicio={fecha_inicio}"
                    + (f" | fecha_demanda={fecha_demanda}" if fecha_demanda else "")
                    + f" | demanda={'sí' if url_demanda else 'no'}")
            except KeyboardInterrupt:
                raise
            except Exception:
                log(f"[Error] Fallo al procesar {num}/{anio}:\n{traceback.format_exc()}")
            finally:
                if driver:
                    driver.quit()

            try:
                db.guardar_expediente(
                    num, anio, caratula, resultado, participantes,
                    jurisdiccion, juzgado, secretaria, "Reintento",
                    fecha_inicio=fecha_inicio,
                    url_demanda=url_demanda or "NINGUNA",
                    fecha_demanda=fecha_demanda,
                    detalle_demanda=detalle_demanda,
                )
            except Exception:
                log(f"[Error] Fallo al guardar en BD {num}/{anio}:\n{traceback.format_exc()}")

            if on_progreso:
                on_progreso(idx, total)

        log("[Exito] Reintento completado. Resultados actualizados en PostgreSQL.")

    except KeyboardInterrupt:
        print("\n")
        log("[Interrumpido] Proceso detenido. Progreso guardado en la BD.")
    except Exception:
        log(f"[Error] Error en el proceso de reintento:\n{traceback.format_exc()}")


if __name__ == "__main__":
    USER = "20286335528"
    PASS = "Federal2025#"

    # --- Modo individual (un expediente) ---
    # ejecutar_flujo(usuario=USER, password=PASS, headless=False)

    # --- Modo batch desde Excel ---
    ejecutar_desde_excel(
        archivo_entrada="expedientes.xlsx",
        usuario=USER,
        password=PASS,
        headless=False
    )
